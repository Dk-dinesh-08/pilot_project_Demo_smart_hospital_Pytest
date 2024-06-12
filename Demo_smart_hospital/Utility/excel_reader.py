import openpyxl

def get_data(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_columns = sheet.max_column
    for r in range(2, total_rows + 1):
        row_list = []
        for c in range(1, total_columns + 1):
            cell_value = sheet.cell(r, c).value
            if isinstance(cell_value, int): 
                row_list.append(str(cell_value))  
            else:
                row_list.append(cell_value)
        final_list.append(row_list)
    return final_list
